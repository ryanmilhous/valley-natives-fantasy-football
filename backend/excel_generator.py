"""
Excel Generator for Fantasy Football Data

Creates a comprehensive Excel spreadsheet with multiple sheets
containing historical league data.
"""

import json
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).parent.parent
PROCESSED_DATA_DIR = BASE_DIR / 'data' / 'processed'
EXPORTS_DIR = BASE_DIR / 'data' / 'exports'
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


class ExcelGenerator:
    """Generate Excel spreadsheet from processed fantasy football data"""

    def __init__(self, processed_data=None):
        if processed_data:
            self.data = processed_data
        else:
            self.data = self.load_processed_data()

        self.output_file = None
        self.writer = None

    def load_processed_data(self):
        """Load processed data from JSON"""
        complete_file = PROCESSED_DATA_DIR / 'complete_data.json'
        if not complete_file.exists():
            raise FileNotFoundError(
                "Processed data not found. Run data_processor.py first."
            )

        with open(complete_file, 'r') as f:
            return json.load(f)

    def create_league_overview_sheet(self):
        """Sheet 1: League Overview"""
        metadata = self.data.get('metadata', {})
        teams = self.data.get('teams', {})
        playoffs = self.data.get('playoffs', [])

        overview_data = []

        # League info
        overview_data.append(['League Name', metadata.get('league_name', 'Unknown')])
        overview_data.append(['Total Seasons', metadata.get('total_seasons', 0)])
        overview_data.append(['First Season', metadata.get('first_season', 'N/A')])
        overview_data.append(['Latest Season', metadata.get('latest_season', 'N/A')])
        overview_data.append(['Total Teams', metadata.get('total_teams', 0)])
        overview_data.append(['Total Matchups Played', metadata.get('total_matchups', 0)])
        overview_data.append([])

        # Championships by team
        overview_data.append(['Championships by Team'])
        overview_data.append(['Team', 'Owner', 'Championships', 'Playoff Appearances'])

        for team_id, team_data in teams.items():
            overview_data.append([
                team_data['current_name'],
                team_data['current_owner'],
                team_data['all_time']['championships'],
                team_data['all_time']['playoff_appearances']
            ])

        df = pd.DataFrame(overview_data)
        df.to_excel(self.writer, sheet_name='League Overview', index=False, header=False)

        print("  ✓ Created League Overview sheet")

    def create_season_standings_sheet(self):
        """Sheet 2: Season Standings"""
        standings = self.data.get('standings', [])

        if not standings:
            print("  ⚠ No standings data available")
            return

        df = pd.DataFrame(standings)

        # Reorder columns
        columns = ['year', 'standing', 'team_name', 'owner', 'wins', 'losses', 'ties',
                   'points_for', 'points_against', 'final_standing', 'playoff_seed']
        df = df[columns]

        # Rename for readability
        df.columns = ['Year', 'Standing', 'Team', 'Owner', 'Wins', 'Losses', 'Ties',
                      'Points For', 'Points Against', 'Final Standing', 'Playoff Seed']

        df.to_excel(self.writer, sheet_name='Season Standings', index=False)

        print("  ✓ Created Season Standings sheet")

    def create_all_matchups_sheet(self):
        """Sheet 3: All Matchups"""
        matchups = self.data.get('matchups', [])

        if not matchups:
            print("  ⚠ No matchup data available")
            return

        df = pd.DataFrame(matchups)

        # Reorder columns
        columns = ['year', 'week', 'is_playoff', 'home_team', 'home_score',
                   'away_team', 'away_score', 'winner', 'point_differential']
        df = df[columns]

        # Rename for readability
        df.columns = ['Season', 'Week', 'Playoff', 'Home Team', 'Home Score',
                      'Away Team', 'Away Score', 'Winner', 'Point Diff']

        df.to_excel(self.writer, sheet_name='All Matchups', index=False)

        print("  ✓ Created All Matchups sheet")

    def create_head_to_head_sheet(self):
        """Sheet 4: Head-to-Head Records Matrix (by Owner)"""
        h2h = self.data.get('head_to_head', {})

        if not h2h:
            print("  ⚠ No head-to-head data available")
            return

        # Get all unique owners
        all_owners = sorted(set(h2h.keys()))

        # Create matrix data
        matrix_data = []
        header = ['Owner'] + all_owners
        matrix_data.append(header)

        for owner1 in all_owners:
            row = [owner1]
            for owner2 in all_owners:
                if owner1 == owner2:
                    row.append('-')
                else:
                    record = h2h.get(owner1, {}).get(owner2, {'wins': 0, 'losses': 0, 'ties': 0})
                    w = record['wins']
                    l = record['losses']
                    t = record['ties']
                    if t > 0:
                        row.append(f"{w}-{l}-{t}")
                    else:
                        row.append(f"{w}-{l}")
            matrix_data.append(row)

        df = pd.DataFrame(matrix_data[1:], columns=matrix_data[0])
        df.to_excel(self.writer, sheet_name='Head-to-Head Records', index=False)

        print("  ✓ Created Head-to-Head Records sheet")

    def create_playoff_history_sheet(self):
        """Sheet 5: Playoff History"""
        playoffs = self.data.get('playoffs', [])

        if not playoffs:
            print("  ⚠ No playoff data available")
            return

        playoff_data = []
        for playoff in playoffs:
            playoff_data.append({
                'Year': playoff['year'],
                'Champion': playoff.get('champion', 'N/A'),
                'Runner-Up': playoff.get('runner_up', 'N/A'),
                'Semifinalists': ', '.join(playoff.get('semifinalists', [])),
                'All Playoff Teams': ', '.join(playoff.get('playoff_teams', []))
            })

        df = pd.DataFrame(playoff_data)
        df.to_excel(self.writer, sheet_name='Playoff History', index=False)

        print("  ✓ Created Playoff History sheet")

    def create_records_sheet(self):
        """Sheet 6: Records & Milestones"""
        records = self.data.get('records', {})

        if not records:
            print("  ⚠ No records data available")
            return

        records_data = []

        # Highest Score
        if records.get('highest_score'):
            hs = records['highest_score']
            records_data.append(['Highest Single Week Score', '', '', ''])
            records_data.append(['Team', 'Score', 'Week', 'Year'])
            records_data.append([
                hs.get('team', 'N/A'),
                hs.get('score', 0),
                f"Week {hs.get('week', 'N/A')}",
                hs.get('year', 'N/A')
            ])
            records_data.append([])

        # Lowest Score
        if records.get('lowest_score'):
            ls = records['lowest_score']
            records_data.append(['Lowest Single Week Score', '', '', ''])
            records_data.append(['Team', 'Score', 'Week', 'Year'])
            records_data.append([
                ls.get('team', 'N/A'),
                ls.get('score', 0),
                f"Week {ls.get('week', 'N/A')}",
                ls.get('year', 'N/A')
            ])
            records_data.append([])

        # Biggest Blowout
        if records.get('biggest_blowout'):
            bb = records['biggest_blowout']
            records_data.append(['Biggest Blowout', '', '', '', ''])
            records_data.append(['Winner', 'Loser', 'Point Diff', 'Week', 'Year'])
            records_data.append([
                bb.get('winner', 'N/A'),
                bb.get('home_team') if bb.get('winner') != bb.get('home_team') else bb.get('away_team'),
                bb.get('point_differential', 0),
                f"Week {bb.get('week', 'N/A')}",
                bb.get('year', 'N/A')
            ])
            records_data.append([])

        # Closest Game
        if records.get('closest_game'):
            cg = records['closest_game']
            records_data.append(['Closest Game (Non-Tie)', '', '', '', ''])
            records_data.append(['Winner', 'Loser', 'Point Diff', 'Week', 'Year'])
            records_data.append([
                cg.get('winner', 'N/A'),
                cg.get('home_team') if cg.get('winner') != cg.get('home_team') else cg.get('away_team'),
                cg.get('point_differential', 0),
                f"Week {cg.get('week', 'N/A')}",
                cg.get('year', 'N/A')
            ])
            records_data.append([])

        # Most Points in a Season
        if records.get('most_points_season'):
            mps = records['most_points_season']
            records_data.append(['Most Points in a Season', '', '', ''])
            records_data.append(['Team', 'Points', 'Record', 'Year'])
            records_data.append([
                mps.get('team_name', 'N/A'),
                mps.get('points_for', 0),
                f"{mps.get('wins', 0)}-{mps.get('losses', 0)}",
                mps.get('year', 'N/A')
            ])
            records_data.append([])

        # Most Wins in a Season
        if records.get('most_wins_season'):
            mws = records['most_wins_season']
            records_data.append(['Most Wins in a Season', '', '', ''])
            records_data.append(['Team', 'Wins', 'Points For', 'Year'])
            records_data.append([
                mws.get('team_name', 'N/A'),
                mws.get('wins', 0),
                mws.get('points_for', 0),
                mws.get('year', 'N/A')
            ])

        df = pd.DataFrame(records_data)
        df.to_excel(self.writer, sheet_name='Records & Milestones', index=False, header=False)

        print("  ✓ Created Records & Milestones sheet")

    def create_team_summary_sheet(self):
        """Sheet 7: All-Time Team Summary"""
        teams = self.data.get('teams', {})

        if not teams:
            print("  ⚠ No team data available")
            return

        team_data = []
        for team_id, team_info in teams.items():
            all_time = team_info.get('all_time', {})
            team_data.append({
                'Team': team_info['current_name'],
                'Owner': team_info['current_owner'],
                'Seasons': team_info['seasons_active'],
                'Wins': all_time.get('wins', 0),
                'Losses': all_time.get('losses', 0),
                'Win %': round(all_time.get('wins', 0) / max(all_time.get('wins', 0) + all_time.get('losses', 0), 1) * 100, 1),
                'Points For': all_time.get('points_for', 0),
                'Points Against': all_time.get('points_against', 0),
                'Championships': all_time.get('championships', 0),
                'Playoff Apps': all_time.get('playoff_appearances', 0)
            })

        df = pd.DataFrame(team_data)
        # Sort by championships, then wins
        df = df.sort_values(by=['Championships', 'Wins'], ascending=False)

        df.to_excel(self.writer, sheet_name='All-Time Team Summary', index=False)

        print("  ✓ Created All-Time Team Summary sheet")

    def style_workbook(self):
        """Apply styling to the Excel workbook"""
        workbook = load_workbook(self.output_file)

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Style each sheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Style header row (first row)
            if sheet.max_row > 0:
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border

        workbook.save(self.output_file)
        print("  ✓ Applied styling to workbook")

    def generate(self, output_filename=None):
        """Generate the complete Excel file"""
        if not output_filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            league_name = self.data.get('metadata', {}).get('league_name', 'league')
            league_name = league_name.replace(' ', '_').lower()
            output_filename = f"{league_name}_history_{timestamp}.xlsx"

        self.output_file = EXPORTS_DIR / output_filename

        print(f"\n=== Generating Excel Spreadsheet ===")
        print(f"Output: {self.output_file}\n")

        # Create Excel writer
        self.writer = pd.ExcelWriter(self.output_file, engine='openpyxl')

        # Create all sheets
        self.create_league_overview_sheet()
        self.create_season_standings_sheet()
        self.create_all_matchups_sheet()
        self.create_head_to_head_sheet()
        self.create_playoff_history_sheet()
        self.create_records_sheet()
        self.create_team_summary_sheet()

        # Save workbook
        self.writer.close()

        # Apply styling
        self.style_workbook()

        print(f"\n✓ Excel file generated successfully!")
        print(f"  Location: {self.output_file}\n")

        return self.output_file


def main():
    """Main entry point for Excel generation"""
    try:
        generator = ExcelGenerator()
        output_file = generator.generate()
        print(f"Open the file at: {output_file}")
    except FileNotFoundError as e:
        print(f"Error: {e}")
        print("Make sure you've run data_extractor.py and data_processor.py first.")
    except Exception as e:
        print(f"Error generating Excel file: {e}")
        raise


if __name__ == '__main__':
    main()
